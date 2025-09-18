import openpyxl

class ExcelUtils:

    @staticmethod
    def get_row_count(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def get_column_count(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_column

    @staticmethod
    def get_cell_data(file_path, sheet_name, row, column):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row, column=column).value

    @staticmethod
    def set_cell_data(file_path, sheet_name, row, column, data):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=column).value = data
        workbook.save(file_path)
